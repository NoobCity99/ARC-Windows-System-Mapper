import csv
import datetime as _dt
import json
import os
from typing import Dict, List

from models import AppEntry, RelatedFile

try:
    from openpyxl import Workbook
except ImportError:  # pragma: no cover - optional dependency
    Workbook = None
from utils import normalize_date, parse_size_mb


CSV_HEADERS = [
    "Name",
    "Group",
    "Version",
    "InstallDate",
    "SizeMB",
    "Publisher",
    "InstallLocation",
    "Website",
]

RELATED_FILE_HEADERS = [
    "AppName",
    "AppVersion",
    "AppPublisher",
    "Path",
    "Type",
    "Source",
    "Confidence",
    "Marked",
]


def export_csv(file_path: str, entries: List[AppEntry]) -> None:
    with open(file_path, "w", newline="", encoding="utf-8") as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow(CSV_HEADERS)
        for app in entries:
            writer.writerow([
                app.name,
                app.group,
                app.version,
                app.install_date,
                "" if app.size_mb is None else str(app.size_mb),
                app.publisher,
                app.install_location,
                app.website,
            ])


def export_related_files_csv(file_path: str, entries: List[AppEntry]) -> None:
    with open(file_path, "w", newline="", encoding="utf-8") as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow(RELATED_FILE_HEADERS)
        for app in entries:
            for related in app.related_files:
                writer.writerow([
                    app.name,
                    app.version,
                    app.publisher,
                    related.path,
                    related.kind,
                    related.source,
                    related.confidence,
                    related.marked,
                ])


def export_csv_with_related(file_path: str, entries: List[AppEntry]) -> str:
    export_csv(file_path, entries)
    base, _ext = os.path.splitext(file_path)
    related_path = f"{base}_related_files.csv"
    export_related_files_csv(related_path, entries)
    return related_path


def export_xlsx(file_path: str, entries: List[AppEntry]) -> None:
    if Workbook is None:
        raise RuntimeError("openpyxl is required for XLSX export. Install it with: pip install openpyxl")
    book = Workbook(write_only=True)
    apps_sheet = book.create_sheet("Apps")
    apps_sheet.append(CSV_HEADERS)
    for app in entries:
        apps_sheet.append([
            app.name,
            app.group,
            app.version,
            app.install_date,
            "" if app.size_mb is None else str(app.size_mb),
            app.publisher,
            app.install_location,
            app.website,
        ])

    related_sheet = book.create_sheet("Related Files")
    related_sheet.append(RELATED_FILE_HEADERS)
    for app in entries:
        for related in app.related_files:
            related_sheet.append([
                app.name,
                app.version,
                app.publisher,
                related.path,
                related.kind,
                related.source,
                related.confidence,
                related.marked,
            ])
    book.save(file_path)


def import_csv(file_path: str) -> List[AppEntry]:
    with open(file_path, "r", newline="", encoding="utf-8") as csv_file:
        reader = csv.DictReader(csv_file)
        if not reader.fieldnames:
            raise ValueError("CSV file has no headers.")
        return _entries_from_rows(reader)


def save_json(file_path: str, entries: List[AppEntry]) -> None:
    payload = {
        "exported_at": _dt.datetime.now().isoformat(timespec="seconds"),
        "apps": [entry.to_dict() for entry in entries],
    }
    with open(file_path, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=2)


def load_json(file_path: str) -> List[AppEntry]:
    with open(file_path, "r", encoding="utf-8") as fh:
        data = json.load(fh)
    items = data.get("apps") if isinstance(data, dict) else data
    if not isinstance(items, list):
        raise ValueError("JSON file does not contain an app list.")
    return _entries_from_json(items)


def _entries_from_rows(reader: csv.DictReader) -> List[AppEntry]:
    header_map: Dict[str, str] = {}
    for header in reader.fieldnames or []:
        normalized = _normalize_header(header)
        header_map[normalized] = header

    def get_value(row: Dict[str, str], key: str) -> str:
        header = header_map.get(key, "")
        if not header:
            return ""
        return str(row.get(header, "") or "").strip()

    entries: List[AppEntry] = []
    for row in reader:
        name = get_value(row, "name")
        if not name:
            continue
        entry = AppEntry(
            name=name,
            group=get_value(row, "group"),
            version=get_value(row, "version"),
            install_date=normalize_date(get_value(row, "installdate")),
            size_mb=parse_size_mb(get_value(row, "sizemb")),
            publisher=get_value(row, "publisher"),
            install_location=get_value(row, "installlocation"),
            website=get_value(row, "website"),
        )
        entries.append(entry)
    return entries


def _entries_from_json(items: List[Dict[str, object]]) -> List[AppEntry]:
    entries: List[AppEntry] = []
    for item in items:
        if not isinstance(item, dict):
            continue
        name = str(item.get("name") or "").strip()
        if not name:
            continue
        related_files = _related_files_from_json(item.get("related_files"))
        entry = AppEntry(
            name=name,
            group=str(item.get("group") or "").strip(),
            version=str(item.get("version") or "").strip(),
            install_date=normalize_date(str(item.get("install_date") or "").strip()),
            size_mb=parse_size_mb(item.get("size_mb")),
            publisher=str(item.get("publisher") or "").strip(),
            install_location=str(item.get("install_location") or "").strip(),
            website=str(item.get("website") or "").strip(),
            related_files=related_files,
        )
        entries.append(entry)
    return entries


def _normalize_header(header: str) -> str:
    return "".join(ch for ch in header.lower() if ch.isalnum())


def _related_files_from_json(raw) -> List[RelatedFile]:
    if not isinstance(raw, list):
        return []
    results: List[RelatedFile] = []
    for item in raw:
        if not isinstance(item, dict):
            continue
        path = str(item.get("path") or "").strip()
        if not path:
            continue
        results.append(
            RelatedFile(
                path=path,
                kind=str(item.get("kind") or "file").strip() or "file",
                source=str(item.get("source") or "").strip(),
                confidence=str(item.get("confidence") or "").strip(),
                marked=str(item.get("marked") or "").strip(),
            )
        )
    return results
